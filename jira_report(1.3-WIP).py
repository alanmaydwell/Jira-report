#!/usr/bin/env python

#Used to access Jira
from jira import JIRA
from jira.exceptions import JIRAError

#Used to create and updatedExcel spreadsheets
import openpyxl
##from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
##from openpyxl.cell import get_column_letter, column_index_from_string

#For password input
import getpass

#Generally usefull
import time

#Used with getattr to read nested attributes from values in string
import functools

#Used to create sub-folder
import os

"""
Extracts details fro Jira
Creates Excel spreadsheet
Includes latest comment in report

0.1 Initial version
0.2 tweaks by Jo
0.3 JiraComm.issue_details() improved. Better mapping of details to issue attributes
Now also uses multi_getattr()

0.4 Previous versions were limited to 50 results due to default maximum
now added maxResults=1000 argument to the issue search in
JiraComm.get_project_issues()

0.5 Modified report to include tab that excludes closed issues.

0.6
Now faster when comment retrieved as no longer has to get each one individually
after the main search.
Found out how to have comments included in response from JIRA.search_issues()
which can be done using fields argument. Modified JiraComm.get_project_issues()
to have optional fields argument.
Added specified fields in JiraComm.report() so comments retrieved in main search
and removed separate comment search as no longer necessary
Modified JiraComm.issue_comments() to have optional get_now flag so
Can now either get fresh comment or (new) extract comment details from
current issue content

0.7
Better integrated the new comment handling. Now included in JiraComm.issue_details()
Removed comments argument from JiraComm.report(). Comments reporting no longer
separate.
Also set different headings background colour for each report spreadsheet
and added auto-filter to each tab.

0.8
Now gets the sprint name (customfield_10003)
JiraComm.report() now also retrieves customfield_10003 which holds the sprint name
(Allbeit in an annoying format)
JiraCom.issue_details now processes customfield_10003, extracts sprint name
from it and adds it to returned details under "Sprint" key
JiraComm.report writes sprint name to graph

0.9 added support for new Type custom field (customfield_11100)
JiraComm.report() now also retrieves customfield_11100 for each bug
 and adds "Type" to spreadsheet.
JiraCom.issue_details now processes customfield_11100 and adds to returns
assocated value under "Type" key

1.0 modifield JiraCommm.get_project_issues to cope with getting issues
in multiple batches. This means no longer limited by 1000 issue limit.

1.1 Modified JiraComm to have self.results_folder attribute to hold
    name of sub folder in which results written. Updated JiraComm.report()
    To create this folder (when not already there) and to write to it.
    Also changed initialisation to use if __name__=="main"

1.2
    Previous versions intended to work with Recentric Project (K008)
    (although should work with other projects but custom fields unlikely to match)
    and recofiguration for other projects required updates in multiple cases.
    Now updated to make project-specific customisation easier.
    Custom field handling used to require updates to both JiraComm.issue_details
    and JiraComm.report. Now changed so mapping of spreadsheet column heading,
    field retrieved, attribute from each field required are defined in a single place:
    self.field_mapping. Ordering also defines column order in spreadsheet.

    Added filename_start argument to JiraComm.report

    Column widths (col_widths) and open_statuses now passed as argument to JiraComm.report().
    This makes reconfiguration more easy. Can now be specified in __main__

    Added JiraComm.issue_field_exam(). Shows fields assoicated with issue.
    Can help identify custom fields.

    Added JiraComm.define_reprocess_fns() and

    Reformated these comments!


1.3
    Major Changes. Report writing now more flexible.
    Results now also held as list of dictionaries (in addtion to list
    of Jira issue objects).

    (1) Better handling of multiple projects
        - Removed self.project attribute from from JiraComm
        now has self.projects which auto records projects available to user but
        individual project choice moved to self.get_project_issues()

        JiraComm.__init__ now
        initialises self.issues
        sets the report filename as self.excel_filename

        JiraComm.get_project_issues()
        -Now has project argument
        -Now has clear_old boolian arg. Clears self.issues when True

        JiraComm.extract_info() method added

        JiraComm.report() will rewrite

        Not changed - self.field_mapping and reprocess_mapping

        Added ExcelSheet class (copied from hub checking script)
"""

def multi_getattr(obj, attr, default = None):
    """
    Asign multiple attributes from string.
    Get a named attribute from an object; multi_getattr(x, 'a.b.c.d') is
    equivalent to x.a.b.c.d. When a default argument is given, it is
    returned when any attribute in the chain doesn't exist; without
    it, an exception is raised when a missing attribute is encountered.

    Taken from http://code.activestate.com/recipes/577346-getattr-with-arbitrary-depth/
    """
    attributes = attr.split(".")
    for i in attributes:
        try:
            obj = getattr(obj, i)
        except AttributeError:
            if default:
                return default
            else:
                #Modified by Alan so no longer raises exception
                ##raise
                pass
    return obj


class JiraComm:
    """
    Connects to LAA Jira
    Can be used to retrieve issue details

    args:
        username - jira username
        password - jira password
        folder - sub folder name for results to be written. Only accepts one level.
        field_mapping - list of 3-value tuples containing field mapping for Jira project
        (optional, will default to K008 settings).
        reprocess_mapping - optional dictionary for mapping results fields with functions that
                            reprocess the result. Key is column name from field_mapping.
                            Value is function from self.reprocess
        excel_file_start - optional beggining of excel filename for results file
                            (end of filename, includes date/time is automatic)
    """
    def __init__(self, username, password, folder="Results",
                field_mapping="", reprocess_mapping={}, excel_file_start="Results"):
        #Jira access parameters
        username = username
        password = password
        options =  {'server': 'https://legalaid.atlassian.net'}
        #Results folder
        self.results_folder = folder
        #Set excel filename
        self.excel_filename = excel_file_start + time.strftime("_%d_%m_%Y(%H.%M.%S).xlsx")
        #Full path to the Excel file
        self.excel_file = os.path.join(self.results_folder, self.excel_filename)
        #create ExcelSheet object using the above
        self.excel = ExcelSheet(filename=self.excel_file, newfile=True, tabrename="Info")

        # Extracted details defined a list of tuples (list of lists would work too)
        #   1st [0] - meaningful name/spreadsheet column heading to give to the item.
        #               should be unique
        #   2nd [1] - the official Jira name of the issue field to be retrieved, e.g. "status".
        #           Used to specify details retrieved.
        #   3rd [2] - the corresponding attribute of the returned issue object
        #           containing the item we actually want, e.g. "fields.status.name".
        #           Used to extract the value we want from returned issue object.

        #If use supplied field mapping if one supplied. Otherwise use K008-=
        if field_mapping:
            self.field_mapping = field_mapping
        else:
            #K008 mapping
            self.field_mapping = [
            ("ID","key","key"),
            ("Issue Type","issuetype","fields.issuetype.name"),
            ("Summary","summary","fields.summary"),
            ("Sprint","customfield_10003","fields.customfield_10003"),
            ("Type","customfield_11100","fields.customfield_11100.value"),
            ("Components","components","fields.components"),
            ("Priority","priority","fields.priority.name"),
            ("Status","status","fields.status.name"),
            ("Assignee","assignee","fields.assignee.displayName"),
            ("Reporter","reporter","fields.reporter.displayName"),
            ("Date Created","created","fields.created"),
            ("Date Updated","updated","fields.updated"),
            ("Resolution","resolution","fields.resolution.name"),
            ("Description","description","fields.description"),
            ("Environment","customfield_10101","fields.customfield_10101.value"),
            ("Severity","customfield_10405","fields.customfield_10405.value"),
            ("Latest Comment","comment","fields.comment.comments")
            ]

        #Holds retrieved Jira issue objects
        self.issues = []

        #Holds results extracted from Jira issue objects (should be easier to handle than self.issues)
        self.extracted_results = []

        #Holds some details of most recent run (time run, project code, project name)
        self.latest_runtime =  ""
        self.latest_proj_code = ""
        self.latest_proj_name = ""

        #Define reprocessing (functions used to transform specified elements)
        self.define_reprocess_fns()
        self.reprocess_mapping = reprocess_mapping

        #Try to access Jira using supplied details
        print ""
        print "* Warnings 'SNIMissingWarning' and 'InsecurePlatformWarning' are usual! *"
        print ""
        try:
            self.jira = JIRA(options=options, basic_auth=(username,password))
        except JIRAError as e:
            #The error text accompanying "Unauthorized (401)" is huge and looks like HTML
            print "Access to Jira failed. Invalid username/password?"
            self.jira = None
            self.projects = {}
        else:
            #find all projects available to user
            self.projects = {proj.key:proj.name for proj in self.jira.projects()}


    def define_reprocess_fns(self):
        """Defines some standard functions for reprocessing results and stores
        them in dictionary, self.reprocess

        Some items returned by Jira are in a format which is inconvenient
        (for example a list, which can't be written to a spreadsheet).
        This method defines a number of standard functions which can be used
        to transform the items into something more convenient. These functions
        are stored in dictionary self.reprocess. If mapped to an item by
        reprocess_mapping they will be automatically applied by self.issue_details()
        """
        self.reprocess = {}

        #Re-arrange Jira date to nicer format (uses date_remformat method)
        self.reprocess["date fix"] = self.date_reformat

        #Originally created to concatenate component names into comma-separated string
        self.reprocess["name concat"] = lambda field: ",".join([e.name for e in field])

        #Latest comment
        ##self.reprocess["latest comment"] = lambda comments: "["+comments[-1].author.displayName + ", "+self.date_reformat(comments[-1].updated) + "] "+comments[-1].body
        def temp(comments):
            #Sometimes comment lacks attributes such as author, so exception handling added to skip these
            try:
                #extract info from the latest comment
                info = "["+comments[-1].author.displayName + ", "+self.date_reformat(comments[-1].updated) + "] "+comments[-1].body
            except Exception as e:
                info = ""
            return info
        self.reprocess["latest comment"] = temp

        #Sprint name
        #bit troublesome as issue.fields.customfield_10003 is a list containing a string, from which we only want a substring
        # [u'com.atlassian.greenhopper.service.sprint.Sprint@1af0af6[id=4,rapidViewId=1,state=ACTIVE,name=K008 - Sprint 1,startDate=2016-02-26T11:59:42.924Z,endDate=2016-03-25T11:59:00.000Z,completeDate=<null>,sequence=4]']
        def temp(sprint_field):
            text = sprint_field[0]
            start = 5 + text.find("name=")
            end = text.find(",", start)
            value = text[start:end]
            return value
        self.reprocess["sprint name"]=temp

    def get_project_issues(self, project, clear_old=True, fields=None, max_results=1000):
        """Get list containing all issues associated with the chosen project.
        Store results in self.issues.
        Search string uses JQL (Jira query lang that is)
        Interface has limit of returning maximum of 1000 issues at a time
        but search will automatically repeat if maximum number exceded.

        Args:
            project - project code, e.g. "K008"
            clear_old (bool) - when true, self.issues cleared before adding new
            fields - optional fields included in search results as comma
            separated string, e.g. "key,summary,priority". If None the default
            set of fields included (which is quite a lot but excludes comments)

            max_results - maximum amount of results to retrieve at a time. Jira
            itself has a limit of 1000. If number of results exceedes maximum
            search will keep repeating until all retrieved.
        """
        if clear_old:
            self.issues = []

        #Default fields
        if not fields:
            fields = ",".join([e[1] for e in self.field_mapping])

        #Record to start retrieving from (used to retrieve results in batches)
        start_at = 0

        #Controls while loop that retrieves results
        keep_going = True

        #Set the project but abandon if user cannot access it.
        if project in self.projects:
            search_string = "project="+project
            print "***",search_string,"***"
        else:
            print "User has no access to project:",project
            keep_going = False

        #Set the runtime, project and project code
        self.latest_runtime =  time.strftime("%d/%m/%Y (%H:%M:%S)")
        self.latest_proj_code = project
        self.latest_proj_name = self.projects[project]

        #Get results from Jira
        while keep_going:
            print "Retrieving issues in range: %i, %i" %(start_at, start_at+max_results)
            try:
                issues = self.jira.search_issues(search_string, fields=fields, maxResults=max_results, startAt=start_at)
            except JIRAError as e:
                print "Jira Error when searching for Project's issues.",self.issues,e
                keep_going = False
                issues = []

            #Add the batch of issues to self.issues
            self.issues = self.issues + issues

            #If length of issues == max_results need to do further search to see if there
            #are any more results. Increase start_at value for next search
            if len(issues) == max_results:
                start_at = start_at + max_results
            #Stop if no further results
            else:
                keep_going = False
                print "Finished retrieving issues. Found:",len(self.issues)

        #Extract details from results and store in list of dictionaries
        self.extract_info()

    def get_issue(self,key):
        """Gets single issue from Jira from supplied key
        args:
            key - issue key, eg "K008-17". Can be for any project available to user
            (not constrained by self.project value)
        Returns:
            if issue found, returns issue object
            if none found, returns None
        """
        try:
            issue = self.jira.issue(key)
        except JIRAError as e:
            print "Jira Error when searching for:",key,e
            issue = None
        return issue

    def issue_comments(self,issue,get_now=True):
        """Gets comments associated with issue
        Args:
            issue - Jira issue object
            get_now - (bool) if True get comment details from Jira connection,
            otherwise rely on details already presenet in supplied issue
        Returns:
            list containing dictionary of details for each comment
        """
        #Holds comments
        comments = []

        #NB issue sometimes lacks fields.comment attribute even when actually
        #present in Jira. Seems to be problem with jira.search_issue
        #Can get from individual issue using JIRA.jira.issue() when absent from main search
        if get_now:
            if not self.jira:
                print "Can't get comments because not connected to Jira."
            else:
                #Retrieve comment now
                temp_issue = self.jira.issue(issue.key, fields='comment')
        #If  get_now false, use supplied issue
        else:
            temp_issue = issue

        if temp_issue:
            #Extract details from comment and store in dictionary for convenient access
            #Some issues might not have issue.fields.comment attribute as
            #might be absent if no comment has recorded
            if "comment" in temp_issue.fields.__dict__.keys():
                for comment in temp_issue.fields.comment.comments:
                    temp_dict = {}
                    temp_dict["Author"] = comment.author.displayName
                    temp_dict["Body"] = comment.body
                    temp_dict["Updated"] = self.date_reformat(comment.updated)
                    comments.append(temp_dict)
        return comments

    def issue_details(self, issue):
        """Extract details associated with Jira issue object and return as
        dictionary.
        If an item is mapped to a function in self.reprocess_mapping, the
        selected value will be automatically transformed by the chosen
        function.
        Args:
            issue - Jira issue object
        Returns:
            dictionary containing details
        """
        #Extract details from the issue and store in dictionary, "details"
        #Mapping taken from self.field_mapping
        #Multi-value items may require extra processing.
        details = {}
        for item in self.field_mapping:
            value =  multi_getattr(issue,item[2])
            ##print "v:",value,"i:",item

            #Reprocess the value if we have a reprocess mapping defined
            #and a value to work with
            if item[0] in self.reprocess_mapping and value:
                #Retrieve the corresponding function
                fn_key = self.reprocess_mapping[item[0]]
                fn = self.reprocess[fn_key]
                #Apply the function to the value
                value = fn(value)

            #Store the value
            details[item[0]] = value

        return details

    def extract_info(self):
        """Extracts details of each retrieved issue and stores in list of
        dictionaries for more convenient handling."""
        self.extracted_results = []
        for issue in self.issues:
            details = self.issue_details(issue)
            self.extracted_results.append(details)

    def report(self, results="", headings="", tab="Results", title="", left_col=1, top_row=1):
        """Writes details of Jira issues to spreadsheet

        Args:
            results: list of Jira results objects to write to spreadsheet.
                Defaults to self.issues
            headings: list of column headings to include, in desired order.
                Need to be values from self.field_mapping. Defaults to all columns
                in original order.
            title - title text (project name and code will be automatically added to this)
            tab: name of spreadsheet tab to write to (will be created if not already present)
            left_col : column number of leftmost boundary of area to write to
            top_row - topmost row to write from
        """
        #Create tab if it's not already present
        if tab not in self.excel.wb.sheetnames:
            self.excel.add_tab(tab)

        #Get column number from letter (not currently used)
        ##column = openpyxl.cell.column_index_from_string(col_letter)

        #Default set of results to all of them
        if not results:
            results = self.extracted_results

        #Default headings
        if not headings:
            headings = [e[0] for e in self.field_mapping]

        #Add title
        title = title+" "+self.latest_proj_code+" - "+self.latest_proj_name+ " Count:"+str(len(results))+ " ["+self.latest_runtime+"]"
        self.excel.cell_set(ws_id=tab, row=top_row, column=left_col, value=title, bold=True)

        #Add headings
        self.excel.table_headings(tab=tab, row=top_row+1, column=left_col, headings=headings)

        #Add row data to all results tab for each issue
        for r, result in enumerate(results):
            #Iterating over headings to preserver column order
            for k, key in enumerate(headings):
                ##print details[key]
                #In case key is invalid, check it's present
                if key in result:
                    value = result[key]
                else:
                    value = ""
                    print key,"not found in results."
                #Change None or [] to empty string
                if type(value) in (None, list):
                    value = ""
                #Write issue to spreadsheet
                self.excel.cell_set(ws_id=tab, row=2+top_row+r,
                            column=left_col+k, value=value, border=True)

    def date_reformat(self,jdate):
        """Dates from Jira are strings such as '2016-03-11T15:32:28.000+0000'
        This method converts date into more human-friendly format, eg:
        11-03-2016 15:32:28
        args:
            jdate - date in string format from jira
        returns:
            date in string format but with components re-arraged.
        """
        return jdate[8:10]+"/"+jdate[5:7]+"/"+jdate[:4]+" "+jdate[11:19]

    def issue_field_exam(self,issue):
        """Lists fields associated with a Jira object
        Can help in identifying jira custome field names
        Args:
            issue - JIRA issue object to be examiend
        """
        print ""
        print "*** Attributes associated with issue",issue,"***"
        #for k,v in issue.__dict__.iteritems():
            #print k
        fields = dir(issue.fields)
        for f in fields:
            print f+": ", getattr(issue.fields,f)

    def show_issues(self):
        """Simple way of printing details of currently retrieved issues.

        Setup for K008
        """
        for i in self.issues:
            print ""
            print "==========="
            print "ID:",i.key
            print "==========="
            print "Issue Type:",i.fields.issuetype.name
            print "Summary:",i.fields.summary
            print "Components:", i.fields.components #list
            print "Priority:",i.fields.priority.name
            print "Status:",i.fields.status.name
            print "Assignee:",i.fields.assignee.displayName
            print "Reporter:",i.fields.reporter.displayName
            print "Date Created:",i.fields.created
            print "Date Updated:",i.fields.updated
            print "Resolution:",i.fields.resolution
            print "Description:",i.fields.description
            print "Environment (custom):",i.fields.customfield_10101.value
            print "Severity (custom):",i.fields.customfield_10405.value


class ExcelSheet:
    def __init__(self,filename = "", newfile = True, tabrename = ""):
        """
        Helps create and update Excel spreadsheet (uses openpyxl)
        Args:
            filename (str) - spreadsheet filename (can be full path)
            newfile (bool) - create new filename if True, otherwise load existing
            tabrename (str) - Optional new name for active tab. This is only tab in a new spreadsheet
        """
        #Create/Open spreadsheet
        self.filename = filename
        if newfile:
            #Create a new spreadsheet
            self.wb = openpyxl.Workbook()
        else:
            #Load existing spreadsheet
            self.wb = openpyxl.load_workbook(filename=filename)

        #Rename the active tab if tabrename set (mainly of value for new spreadsheets to set name of their initial tab)
        if tabrename:
            ws = self.wb.active
            ws.sheet_view.showGridLines = False
            ws.title = tabrename

        #Define some standard formatting settings

        #Make 64 colours
        self.fill_colours = []
        self.colours = []
        cvalues = ['00','55','AA','FF']
        for blue in cvalues:
            for red in cvalues:
                for green in cvalues:
                    colour = "FF"+red+green+blue
                    self.fill_colours.append(openpyxl.styles.PatternFill(start_color=colour,end_color=colour,fill_type='solid'))
                    self.colours.append(colour)

        #Define a cell border style
        self.cell_thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
             right=openpyxl.styles.Side(style='thin'),
             top=openpyxl.styles.Side(style='thin'),
             bottom=openpyxl.styles.Side(style='thin'))

    def show_colours(self,ws_id,trow,lcolumn):
        """
        Adds all fill colours to grid in sreadsheet to show what they look like
        Args:
            ws_id - name (str) or index number (int) of tab to be updated
            trow - topmost row
            lcolumn - leftmost column
        """
        ws = self.select_ws(ws_id)
        for fi, fill in enumerate(self.fill_colours):
            row = trow + fi%16
            column = lcolumn + fi/16
            cell = ws.cell(row=row,column = column)
            cell.value = "Colour "+str(fi)
            cell.fill = self.fill_colours[fi]

    def save(self,newfilename=""):
        """Saves the spreadsheet using either original or replacement filename
        Args:
            newfilename (str) - optional new filename for the save file .
            If not set, self.filename will be used.
        """
        if newfilename:
            self.wb.save(newfilename)
        else:
            self.wb.save(self.filename)

    def add_tab(self,title,last=True):
        """Add tab to spreadsheet
        Args:
            title - tab name
            last (bool) - create in last position when True, otherwise first position
        """
        if last:
            #Create in last position
            ws = self.wb.create_sheet(title=title)
        else:
            #Create in first position
            ws = self.wb.create_sheet(index=0,title=title)
        #Turn off gridlines
        ws.sheet_view.showGridLines = False

    def select_ws(self,id):
        """Select worksheet by either index number or name
        Args: id - either index number  (int) or name (str) of
        worksheet to be selected

        Returns:
            chosen worksheet (if successful) otherwise False
        """
        #Default value
        ws = False

        #Select worksheet by index
        if type(id) is int:
            if id>=0 and id< len(self.wb.worksheets):
                ws = self.wb.worksheets[i]
            else:
                print "ExcelSheet.select_ws() Sheet not selected because index value %i is out of range." % (id)
        #Select worksheet by name
        elif type(id) in (str, unicode):
            if id in self.wb:
                ws = self.wb[id]
                ##ws = self.wb.get_sheet_by_name(id)
            else:
                print "ExcelSheet.select_ws() Sheet not selected because name '%s' not present." % (id)
        return ws

    def cell_set(self, ws_id, row, column, value, bold=False, border=False, colour="FF000000", fi=None):
        """Sets value of cell with optional border or styling
        Args:
            ws_id - name (str) or index number (int) of tab to be updated
            row - row number of cell
            column - column number of cell
            value - value to be entered into cell
            bold - when True make bold
            border - when True add a border
            colour - text colour as string of four 2-digit hex numbers, alpha,red,green,blue (e.g. "FF1122A0")
            fi - index number (0 to 63) of background colour from self.fill_colours
        """
        ws = self.select_ws(ws_id)
        cell = ws.cell(row = row, column = column)
        try:
            cell.value=value
        except Exception as e:
            cell.value="<ERROR WRITING VALUE>"

        #Text colour
        cell.font = openpyxl.styles.Font(color=colour)

        if bold:
            cell.font = openpyxl.styles.Font(bold=True, color=colour)
        if border:
            cell.border = self.cell_thin_border
        if fi:
            cell.fill = self.fill_colours[fi]


    def update_col_widths(self,tab,widths):
        """Adjust multiple column widths
        Args:
            tab - name of tab to be updated
            widths - either
                    (a) a dictionary with columns letters as keys and required widths
                    as values, e.g. {"A":15,"B":9,"C":15}
                    (b) a list or tuple of column widths for column A onwards
                    e.g. [15,9,15]
        """
        ws = self.wb[tab]

        #Adjust widths from dictionary
        if type(widths) is dict:
            for col, width in widths.iteritems():
                ws.column_dimensions[col].width = width
        elif type(widths) in [list,tuple]:
            #Adjust widths from list
            for i, width in enumerate(widths):
                ws.column_dimensions[openpyxl.cell.get_column_letter(i+1)].width = width

    def table_headings(self, tab, row, column, headings, fill=True, filter_on=False):
        """Adds table headings to spreadsheet

        Args:
            tab - tab name
            row - row number for headings
            column - column number of leftmost of the headings
            headings - list containing headings
            fill (bool) - Set backgournd colour when True
            filter_on (bool) - add filter to headings when True (note Excel has limit of one auto_filter per tab)
        """
        ws = self.wb[tab]
        for dc, heading in enumerate(headings):
            cell = ws.cell(row = row, column = column+dc )
            cell.value = heading
            cell.border = self.cell_thin_border
            cell.font = openpyxl.styles.Font(bold=True)
            if fill:
                cell.fill = self.fill_colours[54]

        #Add auto filter to headings (range needs to be sting of "A1:C1" type)
        if filter_on:
            filter_range = openpyxl.cell.get_column_letter(column)+str(row)+":"+openpyxl.cell.get_column_letter(column+dc)+str(row)
            ##print "fr:",filter_range
            ws.auto_filter.ref = filter_range

    def table_values(self,tab,row,column,data):
        """Adds grid of values to spreadsheet
        Args:
            tab - tab name
            row - row number for headings
            column - column number of leftmost of the headings
            data - data to be written as list of lists [row][column]
            e.g. [["A","B,"C"],[1,2,3],[5,6,5]]
        """
        ws = self.wb[tab]
        #Write data to tab
        for dr, rowdata in enumerate(data):
            for dc, colvalue in enumerate(rowdata):
                cell =  ws.cell(row=row+dr, column=column+dc)
                cell.value = colvalue
                cell.border = self.cell_thin_border

    def highlighter(self,ws_id,row_range,col_range,conditions,show_exceptions=False):
        """Highlights spreadsheet cells based on passed conditions
        Args:
            ws_id - name (str) or index number (int) of tab to be updated
            row_range - iterable containing row numbers to covered
            col_range - iterable containing column numbers to be covered
            conditions - list of functions with conditional return values, used to determine whether highlighting applied
            show_exceptions (bool) - when true message for each failed exception printed to console.

        Returns:
            list of integers showing number of times each condition satisfied
            in same order as in conditions list
        """
        #Holds number of times each condition satisfied
        counts = [0 for con in conditions]

        ws = self.select_ws(ws_id)
        for row in row_range:
            for column in col_range:
                for ci,cond in enumerate(conditions):
                    #Exception handling - mainly to keep going when incompatible data types encountered
                    try:
                        if cond(ws,row,column):
                            ##print "Filled",ci,row,column
                            ws.cell(row = row, column = column).fill = self.fill_colours[28]

                            #Increment the count
                            counts[ci] += 1
                            ###Don't bother with further conditions for a cell once one has passed
                            ##break

                    except Exception as e:
                        if show_exceptions:
                            print"Condition exception with condition %i, row %i, col %i: "%(ci,row,column),e[0]
        return counts



#Automatically executes if script is run directly but not if script imported as module
if __name__=="__main__":

    #Username and Password
    username = raw_input("Jira username?")
    password = getpass.getpass("Jira Password?")

    #Jira Project Code
    project = 'K008'

    #Spreadsheet filename start (end automatic, includes date/time)
    filename_start =  'ZZZ_Jira_Bugs'

    #Jira field mapping
    #List of 3-item tuples in desired spreadsheet column order.
    #Each tuple holds 3 strings:
    #(1) Column title to be assigned. Should be unique but otherwise set to what you like.
    #(2) Jira official field name of field to be retrieved
    #(3) Attribute of returned issue object that holds details required
    # required because one field can have mutliple returned value, e.g priority has returned name and number.
    #
    #("Issue Status","status","fields.status.name")
    field_mapping = [
    ("ID","key","key"),
    ("Summary","summary","fields.summary"),
    ("Status","status","fields.status.name"),
    ("Issue Type","issuetype","fields.issuetype.name"),
    ("Sprint","customfield_10003","fields.customfield_10003"),
    ("Type","customfield_11100","fields.customfield_11100.value"),
    ("Components","components","fields.components"),
    ("Priority","priority","fields.priority.name"),
    ("Assignee","assignee","fields.assignee.displayName"),
    ("Reporter","reporter","fields.reporter.displayName"),
    ("Date Created","created","fields.created"),
    ("Date Updated","updated","fields.updated"),
    ("Resolution","resolution","fields.resolution.name"),
    ("Description","description","fields.description"),
    ("Environment","customfield_10101","fields.customfield_10101.value"),
    ("Severity","customfield_10405","fields.customfield_10405.value"),
    ("Latest Comment","comment","fields.comment.comments")
    ]

    # Define reprocessing for results where needed
    # Dictionary holds mapping.
    # Keys are column titles from field mapping (above).
    # Values are keys from JiraComm.reprocess dictionary (definitions in JiraComm.define_reprocess_fns())
    # JiraComm.issue_details applies these automatically
    reprocess_mapping = {
    "Sprint":"sprint name",
    "Components":"name concat",
    "Date Created":"date fix",
    "Date Updated":"date fix",
    "Latest Comment":"latest comment"
    }

    # Statuses of bugs included in "Open Bugs" tab of spreadsheet
    open_statuses = ["New Bug", "Investigate/Fix", "In Test", "Failed"]

    #Column widths for spreadsheet
    col_widths = {"A":9,"B":32,"C":23,"D":11,"E":17,"F":14,"G":14,"H":14,"I":14,"J":14,"K":17,"L":17,"N":64,"O":15,"Q":64}

    #Make connection to Jira
    go = JiraComm(username=username, password=password, folder="Results",
            field_mapping=field_mapping, reprocess_mapping=reprocess_mapping,
            excel_file_start = filename_start)

    #If connection successful, do stuff
    if go.jira:
        #Get all bugs for project
        go.get_project_issues(project='K008')
        #Write "All Bugs"
        #Write all to spreadsheet (defaults to all, so need for results argument)
        go.report(top_row=3, left_col=1, tab="All Bugs", title="All Bugs")
        #Adjust widths
        go.excel.update_col_widths(tab="All Bugs",widths=col_widths)

        #Write "Open Bugs"
        #Find subset of results that are open
        open_bugs = [result  for result in go.extracted_results if result["Status"] in open_statuses]
        #Write open bugs to spreadsheeet
        go.report(results=open_bugs, top_row=3, left_col=1, tab="Open Bugs", title="Open Bugs")
        go.excel.update_col_widths(tab="Open Bugs", widths=col_widths)

        #Open bugs by applications
        #Find components aassociated with open bugs
        components = [result["Components"]  for result in open_bugs]
        components = list(set(components))#only want unique values
        components.sort()
        #Create sepearate table for each one
        row_offset = 0
        #Write results for each component
        for component in components:
            bugs = [result  for result in open_bugs if result["Components"]==component]
            go.report(results=bugs, top_row=3+row_offset, left_col=1, tab="Open Bugs by Item", title="Open Bugs by Item")
            row_offset = row_offset + len(bugs) +3
        go.excel.update_col_widths(tab="Open Bugs by Item", widths=col_widths)

        #Add BAU bugs - comments are stored differently!
        go.get_project_issues(project='DEVTEST')
        go.report(top_row=3, left_col=1, tab="Devtest", title="Devtest")
        go.excel.update_col_widths(tab="Devtest", widths=col_widths)

        #Add some hyperlinks to Info tab
        go.excel.cell_set(ws_id="Info", row=1, column=1, value="Contents", bold=True)
        ws = go.excel.wb["Info"]
        for ti, tab in enumerate(go.excel.wb.worksheets):
            cell =  ws.cell(row=3+ti, column=1)
            cell.value = tab.title
            link = "#'" + tab.title + "'!A1"
            cell.hyperlink = (link)

        #Save spreadsheet
        go.excel.save()
        print "Created: "+go.excel_filename
    else:
        print "*** Access Failed ***"

    print "Finished."

    #Example of getting field info from a known issue. Can help identify custom fields
    ##issue = go.get_issue("K008-410")
    ##go.issue_field_exam(issue)
