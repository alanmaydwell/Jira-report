import openpyxl
import pyodbc
import datetime
import time
import os

#Used to read command-line args (sys.argv)
import sys

#Used to run database checks in parallel
import Queue
import threading

#Used to find host name and IP address of PC
import socket

"""
v0.1 - initial version. Has test data in single spreadsheet tab.
v0.2 - SpreadsheetRun class - expanded to read data from multiple tabs and record results in main control tab
        and new summar tab added
        added Skip column support. Also now only runs query a when database, username and password all have values.
v0.3 - Now optionally supports reading SQL from spreadsheet.
       DbCon class updated so that it doesn't automatically run SQL and automatically close the connection.
       Tweaked a few of its methods.

v0.4 - Connection timeout. Added con_timeout argument and self.con_timeout
       to DbCon class (However it might not have much effect!)

v0.5 - Added new SpreadsheetRun.perform_check() method which places the actual
       database check and spreadsheet update for a single database in one place.
       This is to make move to multi-threaded run easier.
       Updated SpreadsheetRun.save() so summary tab is now the default one
       in the saved spreadsheet.

v0.6 - Multi threaded version. Database checks now run in parallel.
       self.queue added to SpreadsheetRun class. SpreadsheetRun.thread_action()
       processes items from queue using multiple threads.

v0.7 - Realised that writing results to spreadsheet tabs by each thread in v0.6
       could go wrong as relies on SpreadsheetRun.datacols dictionary which only
       holds values for tab currently being examined whereas individual thread
       may need earlier tab values.
       Removed SpreadsheetRun.datacols attribute. Now only local datacols dict
       in SpreadsheetRun.process_tab() method.
       Added new SpreadsheetRun.tab_cols attribute. Which is a dictionary
       with tab name keys and datacols values, so positions for each tab
       now kept.
       Changed SpreadsheetRun.perform_check() to take tab name as argument
       rather than a worksheet object. It now selects the appropriate worksheet
       from the tab name and also gets the self.tab_cols from the tab name too.
       Uses these to write results to the right places in the spreadsheet.

v0.8 - The host name and IP address of PC on which script is run are now
       automatically recorded in the summary tab. This is because we could get
       different results depending on the PC (e.g. because TNSNAME.ORA different)

       Also fixed bug in SpreadsheetRun.process_tab() which would try to
       print username and database for skipped row without having read these values.
       Now reads them.

v0.9 - Added support for optional new spreadsheet column "Condition" used to colour
       result pased on whether it passes or fails.
       SpreadsheetRun.process_tab() now reads condition from spreadsheet condition column.
       SpreadsheetRun.perform_check() evaluates condition when present. Sets fill colour
       based on condition outcome.
       Query result now concatenated with separator

       Changed SpreadsheetRun.perform_check() so that it only converts query result to
       string when more than one value returned. A single-value result is left
       with its original data type and written to spreadsheet as this too.

v1.0 - Added support for optional tab and column spreadsheet rows. When used
       query results are written as table to chosen destination.
       "Result Condition" from spreadsheet used to highlight tabulated results.
       SpreadsheetRun.write_results_table() addded write the table
       Condition (when present) only applied to Result column results.
       Optional "Heading" from spreadsheet also added to destination results.

       DbCon.runsql() now: (i)  captures column headings are stores in self.headings.
       (ii) Records SQL execution date/time self.execution_time.

       SpreadsheetRun.process_tab updated to automatically skip an included tab
       if required spreadsheet columns are not present/found (username, password etc)

       Option to whether to update the master spreadsheet with test results added.
       SpreadsheetRun.update_master used by SpreadsheetRun.save() to determined
       whether this happens.

       SpreadsheetRun.save() modified so results spreadsheets now have
       background colour of top row in each tab set to yellow to make them
       more easily distinguishable from the master spreadsheet.

       Changed automatic column width setting in SpreadsheetRun.write_results()
       to only increase column widths. Now won't reduce from existing width.

v1.1 - Summary now has hyperlinks.
       SpreadsheetRun.__init__() now has  tof (int) to shift results columns to
       right on summary tab to make room for hyperlinks
       SpreadsheetRun.save() now adds hyperlinks to all tabs in column A of
       the summary tab
       Also hyperlink back to summary tab added to A2 in each tab

       SpreadsheetRun.save() now bases the results spreadsheet filename on
       the source spreadsheet filename. Still adds "_results" and date/time
       on end.

        SpreadsheetRun class now has results_file attribute holding the filename
        of the results file. This is set by SpreadsheetRun.save()

       if __name__ == "__main__" now loops of list of spreadsheets, so multiple
       spreadsheets can be inlcuded in a single run. Also prints list of all
       results files created upon conclusion.

v.1.2 - Added exception handling to part of SpreadsheetRun.__init__() that reads
        self.wb. Changed SpreadsheeRun.results_file to SpreadsheeRun.response
        now reports either results file creation or exception message.

        Added suport for command-line arguments as source of spreadsheet filenames
        uses sys.argv. Enables handy .bat file usage to associate spreadsheet(s)
        with script.

        Modified SpreadsheetRun.__init__() to skip None tab names when reading
        details from Run tab. This is to stop pointless warning messages.

v1.3    Added support for writing results to chosen row

        SpreadsheetRun.process_tab() now examines columns up to 15 (O)
        SpreadsheetRun.process_tab() now processes result row value and stores
        in params
        SpreadsheeRun.perform_check() now handles result_row argument
        SpreadsheetRun.write_results_table() how has result_row argument

"""


class DbCon:
    def __init__(self, username, password, database, do_nothing=False, con_timeout=0):
        """
        Create pyodbc database connection.
        If successful, retrieves sysdate.
        Result of query stored in self.results
        Errors held in self.errors
        Args:
            username (str) - username
            password (str) - password
            database (str) - database name
            do_nothing (bool) - don't make connection if true
            con_timeout (int) - timeout time in seconds for making database
                                connection. A value of 0 means the database's
                                default will be used. I don't know if this works
                                might actually be timeout for SQL execution.
        """
        #Holds error messages
        self.errors = []
        #Holds results
        self.results = []
        #Holds column headings
        self.headings = []
        #Connection timeout time
        self.con_timeout = con_timeout
        #Execution time for query as date/time string (updated by self.runsql()
        self.execution_time = ""
        self.database = database

        #Sometimes just want empty self.errors and self.results
        if do_nothing:
            return

        #Construct ODBC connection string
        ##self.constring='Driver={Microsoft ODBC for Oracle};Server='+database+';Uid='+username+';Pwd='+password+';'
        self.constring="Driver={Oracle in Oraclient11g_home1};Dbq="+database+";Uid="+username+";Pwd="+password
        self.open()

    def open(self):
        """Open and test database connection"""
        #Try to make database connection using connection string
        try:
            self.cnxn = pyodbc.connect(self.constring, timeout=self.con_timeout)
        except pyodbc.Error as e:
            self.cnxn = None
            self.errors.append(e[1])

    def close(self):
        """If connection exists, close it"""
        if self.cnxn:
            self.cnxn.close()

    def runsql(self, sql):
        """Execute SQL, retrieve results and print them to console
        Args:
            sql (str) - sql to be executed
        """
        self.execution_time = time.strftime("%d-%b-%Y %H:%M:%S")

        #Don't run if there's no connection
        if not self.cnxn:
            self.errors.append("Can't execute SQL because no connection.")
            return
        #Create cursor and execute SQL;
        cursor = self.cnxn.cursor()
        try:
            cursor.execute(sql)
        except Exception as e:
            self.errors.append("Error on execution:"+e[1])
        else:
            #SQL exececution successful - retrieve results
            try:
                rows = cursor.fetchall()
            except pyodbc.ProgrammingError as e:
                self.errors.append("Error on fetching results:"+e[1])
            else:
                #Capture results
                for row in rows:
                    temp = [col for col in row]
                    self.results.append(temp)
                #Also capture column headings
                self.headings = [d[0] for d in cursor.description]


class SpreadsheetRun:
    def __init__(self,filename="", con_timeout=0):
        """Tries to connect to multiple databases using details in specially
        formatted spreadsheet (database_check.xlsx).
        Success/fail for each recorded in spreadsheet and separate copy of
        spreadsheet saved in results sub-folder.

        Args:
            filename of spreadsheet with test data.
            con_timeout (int) - timeout time in seconds for creating database
                                conection. Value of 0 indicates indefinite wait.
        """
        #Create queue to hold database checks to be handled in multi-thread run
        self.queue = Queue.Queue()

        #Holds info about the response to the test (defaults to "no tests run")
        self.response = filename + "- no tests run"

        #Define a fill colours  for the spreadsheet(colours alpha,r,g,b)
        self.fill_colours=[]
        self.fill_colours.append(openpyxl.styles.PatternFill(start_color='FFFF3333',end_color='FFFF3333',fill_type='solid'))# 0 Red for fail
        self.fill_colours.append(openpyxl.styles.PatternFill(start_color='FFB2FF66',end_color='FFB2FF66',fill_type='solid'))# 1 Green for pass
        self.fill_colours.append(openpyxl.styles.PatternFill(start_color='FF66B2FF',end_color='FF66B2FF',fill_type='solid'))# 2 Blue for headings
        self.fill_colours.append(openpyxl.styles.PatternFill(start_color='FFFFFF99',end_color='FFFFFF99',fill_type='solid'))# 3 Yellow
        self.fill_colours.append(openpyxl.styles.PatternFill(start_color='FFFF8000',end_color='FFFF8000',fill_type='solid'))# 4 Orange
        self.fill_colours.append(openpyxl.styles.PatternFill(start_color='FFCC00CC',end_color='FFCC00CC',fill_type='solid'))# 5 Purple

        #Define a cell border style (used on summary tab)
        self.cell_thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
             right=openpyxl.styles.Side(style='thin'),
             top=openpyxl.styles.Side(style='thin'),
             bottom=openpyxl.styles.Side(style='thin'))

        #Row in each tab where column headings are located. Data starts below this row.
        self.heading_row = 6

        #Try to open spreadsheet
        try:
            self.wb = openpyxl.load_workbook(filename=filename)
        #Give up if fails
        except Exception as e:
            self.response = filename+" - Failed to read: "+e.__doc__
            return

        #Give up if run tab not present
        if "Run" not in self.wb.sheetnames:
            print "Can't proceed as no 'Run' tab in spreadsheet.",filename
            return
        #Select Run tab
        ws = self.wb.get_sheet_by_name("Run")

        #Used to determine wether results are saved to the master spreadsheet
        #self.save will save to master (in addition to report) when self.update_master
        self.update_master = str(ws["D5"].value)
        self.update_master = self.update_master.lower()[:1]

        #Read names of tabs to be included in test run
        tabs_in_run = []
        #Read listed tabs from spreadsheet
        for row in range(5,20):
            tab = ws.cell(row=row, column=2).value
            #Only bother if we have a values
            if tab:
                #Only include if listed name matches actual tab name
                if tab in self.wb.sheetnames:
                    tabs_in_run.append(tab)
                    print "Tab",tab,"added to test run."
                else:
                    print "Tab",tab,"not included because it is not present in",filename

        #Dictionary which will hold the key column positions for each tab in run
        self.tab_cols = {k:"" for k in tabs_in_run}

        #Want blank tab called "Summary" to record summary data
        #If there's already a Summary tab, delete it (easier than deleting its content)
        if "Summary" in self.wb.sheetnames:
            self.wb.remove_sheet(self.wb.get_sheet_by_name("Summary"))
        #Create new summary sheet in first position
        self.summary_tab = self.wb.create_sheet(index=0,title="Summary")
        #Hide gridlines in summary sheet
        self.summary_tab.sheet_view.showGridLines = False
        #Add some details to summary tab
        self.summary_tab["A1"].value="Summary of Database Connection Test Results"
        self.summary_tab["A1"].font = openpyxl.styles.Font(bold=True)
        self.summary_tab["A2"].value=time.strftime("Run start: %d-%b-%Y %H:%M:%S")
        #Add host name and IP address to summary
        hostname = socket.gethostname()
        ipaddr = socket.gethostbyname(hostname)
        self.summary_tab["A3"].value="Run from: "+hostname+" - "+ipaddr

        #Records which tabs have tabulated results
        self.tabulated_results = []

        #Process each tab in tab list (index number, ti, used to set results column in summary tab)

        #tab off-set value for summary results columns - shifts to right.
        #Minimum tof is 1 because column numbers start from 1 but ti starts from 0
        tof = 3
        for ti,tab in enumerate(tabs_in_run):
            print "Processing tab",tab
            #Add tab name to title row in summary tab
            self.summary_tab.cell(row=self.heading_row, column=ti+tof).value = tab
            self.summary_tab.cell(row=self.heading_row, column=ti+tof).border = self.cell_thin_border
            self.summary_tab.cell(row=self.heading_row, column=ti+tof).font = openpyxl.styles.Font(bold=True)
            self.summary_tab.cell(row=self.heading_row, column=ti+tof).fill=self.fill_colours[2]
            #Process queries in tab
            self.process_tab(tab, summary_col=ti+tof, con_timeout=con_timeout)

        #Ensures items below will run only after all items in queue have been processed
        self.queue.join()

        #Save the changes
        results_filename = self.save()
        print "*Finished "+filename+"*"

    def process_tab(self, tab_name, summary_col, con_timeout=0):
        """
        Run tests in specified tab
        Args:
            tab_name (str) - name of tab to be processed (must be present)
            summary_col (int) - column number on summary tab where results will be writen
            con_timeout (int) - timeout time in seconds for making database
                                connection. Value of 0 indicates use default.
        """
        #Select the work
        ws = self.wb.get_sheet_by_name(tab_name)

        #Mapping of data source spreadsheet column headings to parameter keys used by this script
        #Not all headings included as Date/Time and Skip are not stored in params
        headings_to_keys = {"Username":"username",
                            "Password":"password",
                            "Database":"database",
                            "SQL":"sql",
                            "Result Tab":"result_tab",
                            "Result Column":"result_col",
                            "Result Row":"result_row",
                            "Local Condition":"condition",
                            "Result Condition":"r_condition",
                            "Heading":"heading"
                            }

        #Find column positions of expected column headings in supplied tab
        headnames=["Database","Username","Password","Result","Date/Time","Skip","SQL","Result Tab","Result Column","Result Row","Result Condition","Local Condition","Heading"]
        datacols = dict.fromkeys(headnames, -1)
        for col in range(1,16):
            value = str(ws.cell(row=self.heading_row,column=col).value)
            if value in datacols:
                datacols[value] = col

        #See if mandatory required columns found. Record missing columns in bad
        missing_cols = []
        for key in ["Database","Username","Password","Result","Date/Time","Skip"]:
            if datacols[key]==-1:
                print "Column", key,"not found in tab",tab_name
                missing_cols.append(key)

        #Store datacols for this present tab
        self.tab_cols[tab_name] = datacols

        #Read start and end row values
        self.start_row = ws["C3"].value
        self.end_row = ws["C4"].value
        if self.start_row<=self.heading_row:
            self.start_row=self.heading_row+1
        if self.end_row<self.start_row:
            print "End row ("+str(self.end_row)+") is less than start row ("+str(self.start_row)+") !"

        #For each row in chosen range, extract key details from spreadsheet,
        #, run query and update spreadsheet with outcome.

        #Loop over each row in chosen range
        for row in range(self.start_row,self.end_row+1):

            #Only run if (a) Skip is not set & (b) required columns found
            skip = ws.cell(row=row, column = datacols["Skip"]).value
            skip = str(skip).lower()[:1]
            if skip !="y" and not missing_cols:

                #Old definition
                """
                #Store parameters for thread in dictionary
                params = {"username":username,
                        "password":password,
                        "database": database,
                        "sql":sql,
                        "con_timeout":con_timeout,
                        "row":row,
                        "tab_name":tab_name,
                        "summary_col":summary_col,
                        "condition":condition,
                        "result_tab":result_tab,
                        "result_col":result_col
                        }

                """

                #Params for check (not from each row). More added in loop below.
                params = {"con_timeout":con_timeout,
                        "row":row,
                        "tab_name":tab_name,
                        "summary_col":summary_col,
                        }

                #Read row-based param values for present row (sql, username, password ..)
                for column in headings_to_keys:
                    #Can only read if column actually present
                    if datacols[column]!=-1:
                        value = ws.cell(row=row, column=datacols[column]).value
                        if value is None:
                            value=""
                        param_key = headings_to_keys[column]
                        params[param_key]=str(value)

                #Remove any carriage returns from r_condition and replace with spaces
                if "r_condition" in params:
                    params["r_condition"] = params["r_condition"].replace("\n"," ").replace("\r", " ")

                #Add params to queue for multi-thread processing
                self.queue.put(params)

                #Create thread to carry out the check - uses self.thread_action()
                t = threading.Thread(target=self.thread_action)
                t.start()

            #Skipped Row - still add note about skipping to summary page
            else:
                summary_cell = self.summary_tab.cell(row=row, column=summary_col)
                summary_cell.border = self.cell_thin_border#Cell border
                if missing_cols:
                    summary_cell.value = "Skipped. Warning missing column(s): "+",".join(missing_cols)
                else:
                    summary_cell.value = "Skipped."
                    #Only printed if columns not missing, so sure database and username can be read
                    database = ws.cell(row=row, column=datacols["Database"]).value
                    username = ws.cell(row=row, column=datacols["Username"]).value
                    print username,database,"SKIPPED"

    def thread_action(self):
        """Method assigned to each database check thread.
        Called from self.process_tab()
        """
        #Get parameters from queue
        params = self.queue.get()
        #Peform the check
        self.perform_check(**params)
        #Mark task as done
        self.queue.task_done()

    def perform_check(self, username, password, database, sql="SELECT SYSDATE FROM DUAL", con_timeout=0, row=None, tab_name=None, summary_col=1,
                    condition="",result_tab="",result_col="", result_row="", r_condition="",heading=""):
        """Create database connection using specified username, database
        and password. If successful, run SQL. Writes results to spreadsheet.
        Args:
            username - username for database
            password - password for database
            database - database name
            sql - sql to run
            row - row number where login data originated. Used to write back results.
            tab_name - Spreadsheet tab name where login data originaged. Used to write back results.
            summary_col - column number in spreadsheet summary tab to write summary results.
            con_timeout (int) - timeout value for ODBC connection (not sure if actually works)
            condition (str) - optional condition used to determine pass/fail value
            result_tab (str) - optional tab name for writing results
            result_col (str) - optional column letter for writing resuls
            result_row - optional row number of writing results
            r_condition (str) - optional condition applied to data in results table
            heading (str) - optional heading for the query
        """
        #Execute the query using DbCon object if we have username/password/database and row not skipped
        if username and password and database:
            dbcheck = DbCon(username,password,database,con_timeout = con_timeout)
            dbcheck.runsql(sql)
            dbcheck.close()
        else:
            dbcheck = DbCon(username,password,database,do_nothing=True)
            dbcheck.errors.append("Not run because username or password or database value is blank.")

        #Fromat errors for writing to spreadsheet
        error_string = ", ".join(dbcheck.errors)

        #Format query results for writing to spreadsheet.
        #If results a single value just keep it.
        if len(dbcheck.results)==1:
            if len(dbcheck.results[0])==1:
                result = dbcheck.results[0][0]
            #Concatenate results for single row response with more than one column
            else:
                result = ",".join([str(c) for c in dbcheck.results[0]])
        #If we have more than one row. Concatenate them all into a string
        #Reformat returned results from database check as strings
        else:
            result = ""
            for ri, r in enumerate(dbcheck.results):
                #Add carriage return when more than one row.
                if ri>0:
                    result=result+"\n"
                #Concatenate row contents
                temp = ",".join([str(c) for c in r])
                result = result+temp

        #Select the worksheet from tab_name
        ws = self.wb.get_sheet_by_name(tab_name)
        #Get the colum positions for the tab
        datacols = self.tab_cols[tab_name]

        #Write date/time to spreadsheet
        ws.cell(row=row, column=datacols["Date/Time"]).value = datetime.datetime.now()

        #Write result to spreadsheet
        resultcell = ws.cell(row=row, column=datacols["Result"])

        #Default colour index for result cell
        c_index = 1#green background

        #Error result
        if dbcheck.errors:
            print database,username,":",error_string
            #Details for query set tab
            resultcell.value = error_string # Write error messages
            c_index = 0#Red background

        #Non-error result
        else:
            ##print database,username,":",result
            resultcell.value = result # Write results of query

            #If there's a supplied condition, check it and change background colour index based on result
            if condition:
                try:
                    x = result
                    check =  eval(condition)
                    #Set background to orange when check fails (otherwise leave at previous value)
                    if not check:
                        c_index=4
                #Set background to purple if exception raised by check
                except Exception as e:
                    print"Condition", condition,"raised exception with value", result
                    c_index=5

        #Change results cell background colour to index value set in checks above
        resultcell.fill = self.fill_colours[c_index]

        #Update summary tab with summary result - database name with green backround for OK, red for error
        summary_cell = self.summary_tab.cell(row=row, column=summary_col)
        summary_cell.border = self.cell_thin_border#Cell border
        if database:
            summary_cell.value = database+" - "+username+" - "+sql#Write database name
        else:
            summary_cell.value = "<No Name!>"

        #Set summary cell background colour
        summary_cell.fill = self.fill_colours[c_index]#set background colour

        #Write results specified tab and column if specified
        if result_tab and result_col:
            #Default result_row to self.heading_row
            #Also ensure it's an integer
            if not result_row:
                result_row = self.heading_row
            else:
                result_row = int(result_row)
            tl, wh = self.write_results_table(dbcheck=dbcheck, tab=result_tab,
            col_letter=result_col, result_row=result_row, r_condition=r_condition, heading=heading)

            self.tabulated_results.append(result_tab+" ("+result_col+")")

    def write_results_table(self, dbcheck, tab, col_letter, result_row, r_condition="", heading=""):
        """Writes results to SQL query to specified location in spreadsheet
        Can also highlight values on basis of supplied condition (r_condition)

        Args:
            dbcheck - DbCon object with already fetched results
            tab (str) - name of spreadsheet tab to which results are to be written
            col_letter (str) - leftmost column where results table written as column letter e.g. "F"
            result_row (int) - topmost row number where results table written (includes heading row)
            r_condition (str) - optional condition which can be applied to each value using x for
            value and c for column position within query (starting from 1) as Python expression,
            e.g. "c==2 and x > 5"
            heading (str) - optional heading for dispalyed results

        Returns:
            Pair of tuples showing location of results. First top-left corner
            coordinates as col, row as numbers, and second is width, height
            of data. e.g. (1,7), (6,20)
        """
        #Create tab if it's not already present
        if tab not in self.wb.sheetnames:
            self.wb.create_sheet(title=tab)
            ws = self.wb.get_sheet_by_name(tab)
            #Hide the grid
            ws.sheet_view.showGridLines = False
            ws["A1"].value = "Tab added by script on "+time.strftime("%d-%b-%Y %H:%M:%S")
            ws["A1"].font = openpyxl.styles.Font(bold=True)

        #Select the tab
        ws = self.wb.get_sheet_by_name(tab)

        #current row to write too
        row = result_row##=self.heading_row
        #Get column number from letter
        column = openpyxl.cell.column_index_from_string(col_letter)

        #Add title
        cell = ws.cell(row = row-1, column = column)
        cell.value = heading+" ("+dbcheck.database+" "+dbcheck.execution_time+")"
        cell.font = openpyxl.styles.Font(bold=True)

        #Add headings to spreadsheet (with condition on end if included)
        headings = dbcheck.headings[:]

        dc=0#delta column
        for dc, heading in enumerate(headings):
            cell = ws.cell(row = row, column = column+dc )
            cell.value = heading
            cell.border = self.cell_thin_border
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = self.fill_colours[2]
            #Adjust column width based on width of heading - but only make bigger
            ##print openpyxl.cell.get_column_letter(dc+1),len(heading)
            head_width = (len(heading)+2)*1.25
            if  ws.column_dimensions[openpyxl.cell.get_column_letter(column+dc)].width < head_width:
                ws.column_dimensions[openpyxl.cell.get_column_letter(column+dc)].width = head_width

        #Write data to spreadsheet
        row = row+1#shift below the heading row

        #Write errors if we have them
        error_string = ", ".join(dbcheck.errors)
        if error_string:
            ws.cell(row=row, column=column).value=error_string
            ws.cell(row=row, column=column).fill=self.fill_colours[0]

        #Iterate over query results and write them
        for dr, rowdata in enumerate(dbcheck.results):
            for dc, colvalue in enumerate(rowdata):
                cell =  ws.cell(row=row+dr, column=column+dc)
                cell.value = colvalue
                cell.border = self.cell_thin_border

                #If there's a supplied condition, check it and change background colour index based on result
                #This operates at the row level
                if r_condition:
                    try:
                        x = colvalue # value as x for use in condition
                        c = dc+1 #column position with 1 for first column, used in condition
                        check =  eval(r_condition)
                        #Set background to orange when check passes (otherwise leave at previous value)
                        if check:
                            ##print c,":",x
                            cell.fill = self.fill_colours[4]#Orange
                    #Set background to purple if exception raised by check
                    except Exception as e:
                        print"Tabular condition", r_condition,"raised exception with value", colvalue,"Error:", e
                        cell.fill = self.fill_colours[5]#purple
                        check = "Exception"

        #Return location of data (left column, top row), (width, height)
        return (column, result_row+1), (len(dbcheck.headings), len(dbcheck.results))

    def save(self):
        """Saves results at end of test
        """
        print ""

        #Update summary with tabulated results (if any) and hyperlinks to other tabs
        ws = self.wb.get_sheet_by_name("Summary")
        #Add details of tabulated results (if we have any) to summary tab
        if self.tabulated_results:
            ws["A4"].value = "Tabulated Results Recorded: " + ", ".join(self.tabulated_results)
        #Add hyperlinks to left column of summary tab
        ws.cell(row=self.heading_row,column=1).value="Tab Hyperlinks"
        ws.cell(row=self.heading_row,column=1).font = openpyxl.styles.Font(bold=True)
        for ti, tab in enumerate(self.wb.worksheets):
            ws.cell(row=1+ti+self.heading_row,column=1).value=tab.title
            link = "#" + tab.title + "!A1"
            ws.cell(row=1+ti+self.heading_row,column=1).hyperlink = (link)

        #Save changes to main spreadsheet
        if self.update_master=='y':
            try:
                self.wb.save(filename)
            except IOError as e:
                print "***Cannot update main spreadsheet***. Is it open?",e
            else:
                print "Spreasheet,",filename,"updated."

        #Save results to another results spreadsheet
        #Change spreadsheet title on Run tab for this version
        ws = self.wb.get_sheet_by_name("Run")
        ws["A1"].value = "Test Results Spreadsheet"
        ws["A1"].font = openpyxl.styles.Font(bold=True)

        #Sub folder for results - setup if not present
        results_folder=os.path.join(os.getcwd(),"results")
        #Make summary tab the active one
        self.wb.active = 0
        #Create the folder if it doesn't exist
        if not os.path.exists(results_folder):
            os.makedirs(results_folder)
        #Add coloured fill to top row of each tab to make spreadsheet more distinctive from original one.
        #Also add hyperlink to summary tab to cell A2
        fill = openpyxl.styles.PatternFill(start_color='50FFFF00',end_color='50888800',fill_type='solid')
        for ws in self.wb.worksheets:
            #Add "Return to summary" hyperlink (except on summary tab itself)
            if ws.title!="Summary":
                ws.cell(row=2,column=1).hyperlink = ("#Summary!A1")
                ws.cell(row=2,column=1).value = ("Summary Hyperlink")
            #Yellow stripe
            for column in range(1,18):
                ws.cell(row=1,column=column).fill = fill

        #Save results to results file
        #Create filename from source file filename without the extension but with "results" and date/time added
        result_filename = os.path.splitext(filename)[0] + time.strftime("_results_[%Y.%m.%d_%H.%M.%S].xlsx")
        results_file = os.path.join(results_folder,result_filename)
        self.wb.save(results_file)
        print "Results also saved to:",results_file
        self.response = filename + " - Results saved: "+results_file


#Below will only exectute if script run directly.
if __name__ == "__main__":

    #List of spreadsheets to process - can be just one
    ##filenames = ['database_check_only.xlsx', 'excel_sql_runner.xlsx']
    ##filenames = ['database_check_only.xlsx', 'cwa_cdc(rows).xlsx', 'MI_STG_check.xlsx', 'MI_SND_check.xlsx', 'MI_DEV_check.xlsx', 'MI_SYS_check.xlsx', 'MI_SIT_check.xlsx']
    ##filenames = ['excel_sql_runner.xlsx']
    filenames = ['database_check_only.xlsx', 'cwa_cdc(rows).xlsx', 'MI_STG_check.xlsx', 'MI_SND_check.xlsx', 'MI_DEV_check.xlsx', 'MI_SYS_check.xlsx', 'MI_SIT_check.xlsx']

    #Replace filenames with command-line arguments if we have any
    #>1 because first argument "python" doesn't count
    if len(sys.argv)>1:
        filenames = sys.argv[1:]

    #Timeout time in seconds. 0 means use the database's default (not clear whether this actually works)
    con_timeout = 0

    #Holds returned response for each spreadsheet
    responses = []

    #Run the tests from each spreadsheet
    for filename in filenames:
        go = SpreadsheetRun(filename,con_timeout)
        responses.append(go.response)

    #List responses
    print "\n*** Run Finished ***"
    for ir, response in enumerate(responses):
        print "("+str(ir+1)+")",response
